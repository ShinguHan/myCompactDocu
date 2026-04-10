import os
import re
import zipfile
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as XLImage

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "ref", "반출증_Template.xlsx")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

TEMPLATE_COLS = 17   # A~Q
TEMPLATE_ROWS = 81
DATE_ROW = 10
DATE_COL = 11        # K
COMPANY_ROW = 12
COMPANY_COL = 1      # A
ITEM_NAME_ROW = 18
ITEM_NAME_COL = 1    # A
ITEM_QTY_COL = 12    # L
ITEM_UNIT_COL = 10   # J = 단위
DATE_YEAR_ROW = 30   # E30: "2025년 월 일 시 분" 하드코딩 텍스트
DATE_YEAR_COL = 5    # E
MAX_ITEMS = 12       # 행 18~29, 품목 최대 12개
IMAGE_START_ROW = 36


def _copy_template_strip_drawings(src: str, dst: str):
    """템플릿 XLSX를 복사하되 드로잉/미디어 파일은 제거.
    XLSX는 ZIP 구조이므로 raw 레벨에서 xl/drawings, xl/media 항목을 빼고 복사.
    worksheet .rels 와 [Content_Types].xml 에서 drawing 참조도 함께 제거."""
    skip_re = re.compile(r'^xl/(drawings|media)/', re.IGNORECASE)
    drawing_rel_re = re.compile(
        r'<Relationship\b[^>]*relationships/drawing[^>]*/>', re.IGNORECASE
    )
    drawing_ct_re = re.compile(
        r'<Override\b[^>]*/drawings/[^>]*/>', re.IGNORECASE
    )

    with zipfile.ZipFile(src, 'r') as zin, \
         zipfile.ZipFile(dst, 'w', zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            if skip_re.match(info.filename):
                continue  # drawing / media 파일 제외
            data = zin.read(info.filename)
            # worksheet .rels → drawing 관계 제거
            if info.filename.endswith('.rels') and 'worksheets' in info.filename:
                text = data.decode('utf-8')
                text = drawing_rel_re.sub('', text)
                data = text.encode('utf-8')
            # [Content_Types].xml → drawing Override 제거
            elif info.filename == '[Content_Types].xml':
                text = data.decode('utf-8')
                text = drawing_ct_re.sub('', text)
                data = text.encode('utf-8')
            zout.writestr(info, data)


def generate_exit_pass(exit_pass) -> str:
    """
    ExitPass ORM 객체를 받아 반출증 Excel 파일 생성, 파일 경로 반환.
    선택된 품목을 하나의 반출증 페이지(행 18~29)에 모두 기입.
    """
    output_path = os.path.join(
        OUTPUT_DIR, f"반출증_{exit_pass.date}_{exit_pass.company.name}.xlsx"
    )
    _copy_template_strip_drawings(TEMPLATE_PATH, output_path)
    wb = load_workbook(output_path)
    ws = wb["반출증"]

    items = [
        {
            "name": link.transaction.item.report_name or link.transaction.item.name,
            "unit": link.transaction.item.unit,
            "quantity": link.transaction.quantity,
            "amount": link.transaction.total_amount,
            "ledger_number": link.transaction.ledger_number,
        }
        for link in exit_pass.transactions
    ]

    # 관리대장 번호: 거래에 있으면 첫 번째 값 사용, 없으면 반출증 ID 사용 (항상 출력)
    ledger = next(
        (item["ledger_number"] for item in items if item.get("ledger_number") is not None),
        exit_pass.id,
    )

    _fill_header(ws, exit_pass.date, exit_pass.company.name, ledger)

    for i, item in enumerate(items[:MAX_ITEMS]):
        row = ITEM_NAME_ROW + i
        ws.cell(row=row, column=ITEM_NAME_COL).value = item["name"]
        # "원/kg" → "kg", "원/EA" → "EA" 형식 변환
        raw_unit = item.get("unit") or ""
        physical_unit = raw_unit.split("/")[-1] if "/" in raw_unit else raw_unit
        ws.cell(row=row, column=ITEM_UNIT_COL).value = physical_unit
        ws.cell(row=row, column=ITEM_QTY_COL).value = item["quantity"]

    _set_print_layout(ws, 1)

    # 사진 삽입
    if exit_pass.photo_path and os.path.exists(exit_pass.photo_path):
        _insert_photo(ws, exit_pass.photo_path)

    wb.save(output_path)
    _fix_sheet_format(output_path)
    return output_path


def _fill_header(ws, tx_date, company_name, ledger):
    """헤더 영역(C8 관리대장번호, K10 반출일, A12 업체명, E30 반출년월일) 채우기."""
    from openpyxl.styles import Font

    # C8: 관리대장 번호 (4자리 0패딩, 항상 출력)
    c8 = ws.cell(row=8, column=3)
    c8.value = f"{int(ledger):04d}"
    c8.font = Font(size=11, name=c8.font.name, bold=c8.font.bold)

    # K10: 반출 날짜
    ws.cell(row=DATE_ROW, column=DATE_COL).value = tx_date

    # A12: 업체명
    ws.cell(row=COMPANY_ROW, column=COMPANY_COL).value = company_name

    # E30: "2 0 2 6 년     월    일    시    분" 형식으로 연도 갱신
    year = tx_date.year if hasattr(tx_date, 'year') else int(str(tx_date)[:4])
    year_spaced = ' '.join(str(year))
    ws.cell(row=DATE_YEAR_ROW, column=DATE_YEAR_COL).value = (
        f"{year_spaced} 년     월    일    시    분"
    )


def _get_template_area_emu() -> tuple[int, int]:
    """템플릿 XML에서 사진 영역 가로(A~Q) · 세로(row36~81) EMU 계산."""
    import re
    EMU_PER_PT = 12700
    MDW_PX = 7        # Calibri 11pt MaxDigitWidth
    EMU_PER_PX = 9525

    total_w_emu = 0
    total_h_emu = 0

    with zipfile.ZipFile(TEMPLATE_PATH) as z:
        for name in z.namelist():
            if 'xl/worksheets/sheet1' in name and name.endswith('.xml'):
                xml = z.read(name).decode('utf-8')

                # <cols> 파싱 → 컬럼별 width(chars)
                col_widths: dict[int, float] = {}
                cols_m = re.search(r'<cols>(.*?)</cols>', xml, re.DOTALL)
                if cols_m:
                    for entry in re.finditer(
                        r'<col\b[^>]*min="(\d+)"[^>]*max="(\d+)"[^>]*width="([^"]+)"',
                        cols_m.group(1)
                    ):
                        for c in range(int(entry.group(1)), int(entry.group(2)) + 1):
                            col_widths[c] = float(entry.group(3))

                for c in range(1, TEMPLATE_COLS + 1):
                    w_chars = col_widths.get(c, 8.43)
                    total_w_emu += int(w_chars * MDW_PX + 5) * EMU_PER_PX

                # <row> 파싱 → row36~81 높이(pt)
                # ' ht=' 공백 prefix로 customHeight="1" 오매칭 방지
                row_heights: dict[int, float] = {}
                for entry in re.finditer(r'<row\b[^>]* r="(\d+)"[^>]* ht="([\d.]+)"', xml):
                    r = int(entry.group(1))
                    if IMAGE_START_ROW <= r <= TEMPLATE_ROWS:
                        row_heights[r] = float(entry.group(2))

                default_h_pt = 15.0
                m = re.search(r'defaultRowHeight="([^"]+)"', xml)
                if m:
                    default_h_pt = float(m.group(1))

                for r in range(IMAGE_START_ROW, TEMPLATE_ROWS + 1):
                    total_h_emu += int(row_heights.get(r, default_h_pt) * EMU_PER_PT)
                break

    return total_w_emu, total_h_emu


def _insert_photo(ws, photo_path: str):
    """반출증 사진 영역(row 36~81)에 이미지 삽입. 영역에 맞게 스케일, 가로 중앙 정렬."""
    try:
        from PIL import Image as PILImage
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D

        # PIL로 실제 픽셀 크기 읽기
        with PILImage.open(photo_path) as pil_img:
            orig_w, orig_h = pil_img.size

        # 사진 영역 크기(EMU) — 템플릿 XML 기반, 5% 여백
        area_w_emu, area_h_emu = _get_template_area_emu()
        usable_w = int(area_w_emu * 0.85)
        usable_h = int(area_h_emu * 0.90)

        # 비율 유지하며 영역에 맞게 스케일
        scale = min(usable_w / orig_w, usable_h / orig_h)
        img_w_emu = int(orig_w * scale)
        img_h_emu = int(orig_h * scale)

        # 상하좌우 중앙 정렬
        x_emu = max(0, (area_w_emu - img_w_emu) // 2)
        y_emu = max(0, (area_h_emu - img_h_emu) // 2)

        img = XLImage(photo_path)
        img.anchor = OneCellAnchor(
            _from=AnchorMarker(col=0, colOff=x_emu, row=IMAGE_START_ROW - 1, rowOff=y_emu),
            ext=XDRPositiveSize2D(img_w_emu, img_h_emu)
        )
        ws.add_image(img)
    except Exception:
        pass


def _set_print_layout(ws, item_count):
    from openpyxl.worksheet.pagebreak import Break
    last_col = get_column_letter(item_count * TEMPLATE_COLS)
    ws.print_area = f"$A$1:${last_col}${TEMPLATE_ROWS}"
    for i in range(1, item_count):
        ws.col_breaks.append(Break(id=i * TEMPLATE_COLS))


def _copy_template_block(ws, col_offset):
    for row in range(1, TEMPLATE_ROWS + 1):
        for col in range(1, TEMPLATE_COLS + 1):
            src = ws.cell(row=row, column=col)
            dst = ws.cell(row=row, column=col + col_offset)

            if src.value is not None:
                if isinstance(src.value, str) and src.value.startswith("="):
                    dst.value = _shift_formula(src.value, col_offset)
                else:
                    dst.value = src.value

            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format

    new_merges = []
    for mr in list(ws.merged_cells.ranges):
        if (1 <= mr.min_col <= TEMPLATE_COLS and 1 <= mr.max_col <= TEMPLATE_COLS
                and 1 <= mr.min_row <= TEMPLATE_ROWS and 1 <= mr.max_row <= TEMPLATE_ROWS):
            min_col_l = get_column_letter(mr.min_col + col_offset)
            max_col_l = get_column_letter(mr.max_col + col_offset)
            new_merges.append(f"{min_col_l}{mr.min_row}:{max_col_l}{mr.max_row}")
    for m in new_merges:
        ws.merge_cells(m)

    for key, src_dim in list(ws.column_dimensions.items()):
        min_col = src_dim.min or column_index_from_string(key)
        max_col = src_dim.max or column_index_from_string(key)
        if min_col > TEMPLATE_COLS or max_col < 1:
            continue
        for col in range(max(min_col, 1), min(max_col, TEMPLATE_COLS) + 1):
            ws.column_dimensions[get_column_letter(col + col_offset)].width = src_dim.width


def _fix_sheet_format(path: str):
    """openpyxl 저장 시 변형되는 컬럼 너비(<cols>)를 템플릿 원본으로 복원하고
    baseColWidth 속성도 제거."""
    import re

    # 템플릿에서 <cols>...</cols> 블록 추출
    template_cols_xml = None
    with zipfile.ZipFile(TEMPLATE_PATH, 'r') as zt:
        for name in zt.namelist():
            if 'xl/worksheets/sheet' in name and name.endswith('.xml'):
                xml = zt.read(name).decode('utf-8')
                m = re.search(r'<cols>.*?</cols>', xml, re.DOTALL)
                if m:
                    template_cols_xml = m.group()
                break

    all_files = {}
    with zipfile.ZipFile(path, 'r') as z:
        for name in z.namelist():
            all_files[name] = z.read(name)

    changed = False
    for name, data in all_files.items():
        if 'xl/worksheets/sheet' in name and name.endswith('.xml'):
            text = data.decode('utf-8')
            new_text = re.sub(r'\s*baseColWidth="\d+"', '', text)
            # 반출증 시트(sheet1)만 <cols> 복원
            if template_cols_xml and name.endswith('sheet1.xml'):
                new_text = re.sub(r'<cols>.*?</cols>', template_cols_xml, new_text, flags=re.DOTALL)
            if new_text != text:
                all_files[name] = new_text.encode('utf-8')
                changed = True

    if changed:
        tmp = path + '._tmp'
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name, data in all_files.items():
                zout.writestr(name, data)
        os.replace(tmp, path)


def _shift_formula(formula, col_offset):
    if "#REF!" in formula:
        return formula
    ext_ref_pattern = r"('[^']+'![^\s,();]+|[A-Za-z_]\w*![^\s,();]+)"
    parts = re.split(ext_ref_pattern, formula)
    result = []
    for part in parts:
        if "!" in part:
            result.append(part)
        else:
            result.append(_shift_internal_refs(part, col_offset))
    return "".join(result)


def _shift_internal_refs(text, col_offset):
    def replacer(match):
        dollar1, col_str, dollar2, row_str = match.groups()
        try:
            new_col = get_column_letter(column_index_from_string(col_str) + col_offset)
            return dollar1 + new_col + dollar2 + row_str
        except Exception:
            return match.group(0)

    return re.sub(r"(?<![!\w])(\$?)([A-Z]{1,3})(\$?)(\d+)", replacer, text)
