"""
엑셀 원본 파일에서 데이터 임포트 서비스.
'폐기물, 부산물 입출고대장' 시트를 파싱해서 DB에 저장.
"""
import io
from typing import List
from datetime import date

import openpyxl
from sqlalchemy.orm import Session

import models
import schemas


def parse_excel_preview(file_obj: io.BytesIO, db: Session) -> schemas.ImportPreview:
    """엑셀 파일 파싱 후 미리보기 (dry-run)"""
    wb = openpyxl.load_workbook(file_obj, data_only=True)

    # 시트 이름 찾기
    target_sheet = None
    for name in wb.sheetnames:
        if "입출고대장" in name:
            target_sheet = wb[name]
            break
    if not target_sheet:
        raise ValueError("'입출고대장' 시트를 찾을 수 없습니다")

    # 기존 (date, item_name, company_name) 셋으로 중복 체크
    existing = set()
    for tx in db.query(models.Transaction).options(
        __import__("sqlalchemy.orm", fromlist=["joinedload"]).joinedload(models.Transaction.item),
        __import__("sqlalchemy.orm", fromlist=["joinedload"]).joinedload(models.Transaction.company),
    ).all():
        existing.add((tx.date, tx.item.name, tx.company.name))

    # 기존 품목명 셋
    known_items = {item.name for item in db.query(models.Item).all()}

    rows: List[schemas.ImportPreviewRow] = []
    unknown_items = set()

    # row2가 헤더: 일자, 품목, 처리업체, 데이터변환, 계약단가, 단위, 처리량, 금액, 비고
    for row in target_sheet.iter_rows(min_row=3, values_only=True):
        tx_date = row[1]
        item_name = row[2]
        company_name = row[3]
        unit_price = row[5]
        quantity = row[7]
        total_amount = row[8]
        note = row[9]

        if not tx_date or not item_name or not company_name:
            continue
        if hasattr(tx_date, "date"):
            tx_date = tx_date.date()

        if item_name not in known_items:
            unknown_items.add(item_name)

        is_dup = (tx_date, item_name, company_name) in existing

        rows.append(schemas.ImportPreviewRow(
            date=tx_date,
            item_name=str(item_name),
            company_name=str(company_name),
            quantity=float(quantity or 0),
            unit_price=float(unit_price or 0),
            total_amount=float(total_amount or 0),
            note=str(note) if note else None,
            is_duplicate=is_dup,
        ))

    new_count = sum(1 for r in rows if not r.is_duplicate)
    dup_count = len(rows) - new_count

    return schemas.ImportPreview(
        new_count=new_count,
        duplicate_count=dup_count,
        unknown_items=list(unknown_items),
        rows=rows,
    )


def confirm_import(rows: List[schemas.TransactionCreate], db: Session) -> List[models.Transaction]:
    """미리보기 확인 후 실제 DB 저장"""
    from sqlalchemy.orm import joinedload

    saved = []
    latest = (
        db.query(models.Transaction)
        .filter(models.Transaction.ledger_number.is_not(None))
        .order_by(models.Transaction.ledger_number.desc(), models.Transaction.id.desc())
        .first()
    )
    next_ledger = (latest.ledger_number if latest and latest.ledger_number is not None else 0) + 1

    for r in rows:
        data = r.model_dump()
        if data.get("ledger_number") is None:
            data["ledger_number"] = next_ledger
            next_ledger += 1
        tx = models.Transaction(**data)
        db.add(tx)
        saved.append(tx)
    db.commit()

    ids = [tx.id for tx in saved]
    return (
        db.query(models.Transaction)
        .options(
            joinedload(models.Transaction.item),
            joinedload(models.Transaction.company),
        )
        .filter(models.Transaction.id.in_(ids))
        .all()
    )
