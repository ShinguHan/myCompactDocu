"""월말 결산보고서 Excel 생성 서비스

핵심 전략: openpyxl 로 템플릿 구조를 READ-ONLY 스캔한 뒤,
셀 값은 xlsx zip 내부 XML 을 직접 패치.
→ styles.xml·chart XML 등 서식 파일을 절대 건드리지 않으므로
  테두리 변형·차트 손상·Excel 복구 오류가 원천적으로 없음.
"""
import os
import re
import shutil
import zipfile
from openpyxl import load_workbook

import schemas

# ── 경로 ──────────────────────────────────────────────────────────────────────
_TEMPLATE = os.path.join(os.path.dirname(__file__), '..', 'templates', 'report_template.xlsx')
_FALLBACK  = os.path.join(os.path.dirname(__file__), '..', '..', 'ref',
                           '(찐) 26년_3월_부산물 매각 및 폐기물 처리현황.xlsx')
SOURCE_PATH = _TEMPLATE if os.path.exists(_TEMPLATE) else _FALLBACK
OUTPUT_DIR  = os.path.join(os.path.dirname(__file__), '..', '..', 'output')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── 열 인덱스 (1-based) ───────────────────────────────────────────────────────
COL_COMPANY = 2    # B  업체명
COL_ITEM    = 6    # F  품명
COL_UNIT    = 10   # J  단가
COL_QTY     = 12   # L  당월 수량
COL_AMT     = 14   # N  당월 금액
COL_PREV    = 17   # Q  전월 금액
COL_NOTE    = 20   # T  비고

# 요약 섹션 열
COL_S_PQTY      = 4    # D  전월 수량
COL_S_PAMT      = 6    # F  전월 금액
COL_S_QTY       = 9    # I  당월 수량
COL_S_AMT       = 11   # K  당월 금액
COL_S_PREV_AVG  = 14   # N  전년도 월평균
COL_S_AVG       = 17   # Q  당해연도 월평균
COL_S_CUM       = 20   # T  누계

# 섹션 행 범위
BY_DATA_START  = 16
BY_TOTAL_ROW   = 37
WS_DATA_START  = 42
WS_TOTAL_ROW   = 56

# 요약 행
ROW_BY = 8
ROW_WS = 9
ROW_NE = 10

# 초기화 대상 열
CLEAR_SUMMARY_COLS = [4, 6, 9, 11, 14, 17, 20]
CLEAR_DETAIL_COLS  = [12, 14, 17, 20]


# ── XML 직접 패치 유틸 ────────────────────────────────────────────────────────

def _col_letter(n: int) -> str:
    """1-based 열 번호 → Excel 열 문자 (1→A, 107→DC 등)"""
    s = ''
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _addr(row: int, col: int) -> str:
    return f'{_col_letter(col)}{row}'


def _set_cell(xml: str, addr: str, value) -> str:
    """워크시트 XML 문자열에서 특정 셀 값을 설정·초기화.

    value:
      None        → 셀 내용 비움 (자기닫힘 태그로 변환)
      int/float   → 숫자 값 <v>
      str         → 인라인 문자열 t="inlineStr"

    처리 패턴:
      <c r="L8" s="5"><v>123</v></c>                  → 값 교체
      <c r="I8" s="7"><f>L37</f><v>47</v></c>         → 수식 제거 후 값 교체
      <c r="DC25" s="3"/>                              → 확장 후 값 삽입
      <c r="B8" s="2" t="s"><v>20</v></c>             → t 제거 후 값 교체
    """
    esc = re.escape(addr)

    # r="ADDR" 위치 찾기
    ref_m = re.search(rf'\br="{esc}"', xml)
    if not ref_m:
        return xml  # 셀 없음 → 스킵

    # 해당 <c 요소 시작 위치
    c_start = xml.rfind('<c', 0, ref_m.start())
    if c_start == -1:
        return xml

    tail = xml[c_start:]

    # 자기닫힘 여부 판별
    sc_m = re.match(r'<c\b[^>]*/>', tail)
    if sc_m:
        c_end = c_start + sc_m.end()
        old_cell = sc_m.group(0)
        # 여는 태그: '/>' 제거
        open_tag = old_cell[:-2]
        is_sc = True
    else:
        gt = tail.index('>')
        cl = tail.index('</c>', gt)
        c_end = c_start + cl + 4
        old_cell = tail[:cl + 4]
        open_tag = old_cell[:old_cell.index('>')]
        is_sc = False

    # t="..." 속성 제거
    open_tag = re.sub(r'\s+t="[^"]*"', '', open_tag).rstrip()

    if value is None:
        if is_sc:
            return xml  # 이미 비어 있음
        new_cell = open_tag + '/>'
        return xml[:c_start] + new_cell + xml[c_end:]

    if isinstance(value, str):
        esc_v = (value.replace('&', '&amp;')
                      .replace('<', '&lt;')
                      .replace('>', '&gt;'))
        new_cell = f'{open_tag} t="inlineStr"><is><t>{esc_v}</t></is></c>'
    else:
        v = int(value) if isinstance(value, float) and value == int(value) else value
        new_cell = f'{open_tag}><v>{v}</v></c>'

    return xml[:c_start] + new_cell + xml[c_end:]


def _patch_xlsx(xlsx_path: str, sheet_updates: dict):
    """xlsx zip 내 워크시트 XML만 직접 수정. 다른 파일은 일절 건드리지 않음.

    sheet_updates: {'xl/worksheets/sheet1.xml': {(row, col): value, ...}, ...}
    """
    all_files: dict[str, bytes] = {}
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        for name in z.namelist():
            all_files[name] = z.read(name)

    for sheet_path, updates in sheet_updates.items():
        if sheet_path not in all_files:
            continue
        xml_str = all_files[sheet_path].decode('utf-8')
        for (row, col), value in updates.items():
            xml_str = _set_cell(xml_str, _addr(row, col), value)
        all_files[sheet_path] = xml_str.encode('utf-8')

    # calcChain.xml 제거: 수식 셀을 값으로 교체했을 때 Excel이
    # "복구 오류"를 내는 원인 → 삭제하면 Excel이 자동 재계산
    all_files.pop('xl/calcChain.xml', None)

    tmp = xlsx_path + '._tmp'
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as z:
        for name, data in all_files.items():
            z.writestr(name, data)
    os.replace(tmp, xlsx_path)


# ── openpyxl READ-ONLY 스캔 유틸 ─────────────────────────────────────────────

def _normalize(name) -> str:
    """공백·개행·언더스코어·괄호·슬래시 등을 제거해 비교용 키 생성"""
    if not name:
        return ''
    s = str(name)
    s = s.split('/')[0]
    s = re.sub(r'\s+', '', s)
    s = s.replace('\xa0', '').replace('\u00a0', '')
    s = s.replace('_', '')
    s = s.replace('(', '').replace(')', '')
    return s.lower()


def _scan_section(ws, data_start: int, total_row: int):
    """템플릿 워크시트 스캔: 행 번호 매핑 반환 (read-only 전용)"""
    mapping   = {}
    subtotals = {}
    current_cn = None

    for r in range(data_start, total_row + 1):
        b_val = ws.cell(r, COL_COMPANY).value
        f_val = ws.cell(r, COL_ITEM).value
        f_str = _normalize(f_val)

        if b_val is not None:
            b_str = _normalize(b_val)
            if b_str and b_str not in ('합계', '소계'):
                current_cn = b_str

        if r == total_row:
            continue

        if '소계' in f_str:
            if current_cn:
                subtotals[current_cn] = r
        elif f_str:
            mapping[(current_cn, f_str)] = r

    return mapping, subtotals


def _fill_section_updates(updates: dict, rows, mapping: dict,
                           subtotals: dict, total_row: int):
    """섹션 데이터를 updates 딕셔너리에 반영"""
    company_acc: dict = {}

    for r in rows:
        cn  = _normalize(r.company_name)
        itm = _normalize(r.item_name)
        key = (cn, itm)

        if key in mapping:
            row_num = mapping[key]
            updates[(row_num, COL_UNIT)] = r.unit_price or None
            updates[(row_num, COL_QTY)]  = r.current_quantity or None
            updates[(row_num, COL_AMT)]  = r.current_amount or None
            updates[(row_num, COL_PREV)] = r.prev_amount or None
            if r.note:
                updates[(row_num, COL_NOTE)] = r.note

        if cn not in company_acc:
            company_acc[cn] = {'qty': 0, 'amt': 0, 'prev': 0}
        company_acc[cn]['qty']  += r.current_quantity or 0
        company_acc[cn]['amt']  += r.current_amount
        company_acc[cn]['prev'] += r.prev_amount

    for cn, st_row in subtotals.items():
        if cn in company_acc:
            t = company_acc[cn]
            updates[(st_row, COL_QTY)]  = t['qty'] or None
            updates[(st_row, COL_AMT)]  = t['amt'] or None
            updates[(st_row, COL_PREV)] = t['prev'] or None

    total_qty  = sum(r.current_quantity or 0 for r in rows)
    total_amt  = sum(r.current_amount for r in rows)
    total_prev = sum(r.prev_amount for r in rows)
    updates[(total_row, COL_QTY)]  = total_qty or None
    updates[(total_row, COL_AMT)]  = total_amt or None
    updates[(total_row, COL_PREV)] = total_prev or None


# ── 보고서 생성 ───────────────────────────────────────────────────────────────

def generate_monthly_report(summary: schemas.MonthlySummary,
                             trend_data=None) -> str:
    year, month = summary.year, summary.month
    output_path = os.path.join(OUTPUT_DIR,
                               f'월말보고서_{year}년{month:02d}월.xlsx')

    # ── 1. 템플릿 복사 ────────────────────────────────────────────────────────
    shutil.copy(SOURCE_PATH, output_path)

    # ── 2. 템플릿 구조 스캔 (READ-ONLY, 수정 없음) ─────────────────────────
    wb_ro = load_workbook(SOURCE_PATH, data_only=True)
    ws_ro  = wb_ro.worksheets[0]
    cws_ro = wb_ro.worksheets[1]

    # 그래프 시트 row 23에서 당해연도 전체 월→컬럼 매핑 구성
    yy = year % 100
    year_col_map: dict[int, int] = {}   # { month: col_num }
    for cell in cws_ro[23]:
        if cell.value is None:
            continue
        cv = str(cell.value).strip().lstrip("'")
        m = re.match(rf'^{yy:02d}\.(\d+)월$', cv)
        if m:
            year_col_map[int(m.group(1))] = cell.column

    by_mapping, by_subtotals = _scan_section(ws_ro, BY_DATA_START, BY_TOTAL_ROW)
    ws_mapping, ws_subtotals = _scan_section(ws_ro, WS_DATA_START, WS_TOTAL_ROW)
    wb_ro.close()

    # ── 3. 셀 업데이트 딕셔너리 구성 ─────────────────────────────────────────
    s1: dict = {}   # sheet1 (보고서)
    s2: dict = {}   # sheet2 (그래프)

    # ── 3-1. 요약 셀 초기화 ──────────────────────────────────────────────────
    for r in [ROW_BY, ROW_WS, ROW_NE]:
        for c in CLEAR_SUMMARY_COLS:
            s1[(r, c)] = None

    # ── 3-2. 제목 (연도·월) ───────────────────────────────────────────────────
    s1[(2, 2)] = year
    s1[(2, 7)] = month

    # ── 3-3. 연도 레이블 ─────────────────────────────────────────────────────
    s1[(5, COL_S_PREV_AVG)] = year - 1
    s1[(5, COL_S_AVG)]      = year

    # ── 3-4. 요약 섹션 값 ────────────────────────────────────────────────────
    by_curr_qty = sum(r.current_quantity or 0 for r in summary.byproducts)
    ws_curr_qty = sum(r.current_quantity or 0 for r in summary.wastes)

    s1[(ROW_BY, COL_S_PQTY)]     = summary.total_prev_byproduct_qty or None
    s1[(ROW_BY, COL_S_PAMT)]     = summary.total_prev_byproduct or None
    s1[(ROW_BY, COL_S_QTY)]      = by_curr_qty or None
    s1[(ROW_BY, COL_S_AMT)]      = summary.total_current_byproduct or None
    s1[(ROW_BY, COL_S_PREV_AVG)] = (round(summary.prev_year_avg_byproduct)
                                    if summary.prev_year_avg_byproduct else None)
    s1[(ROW_BY, COL_S_AVG)]      = round(summary.ytd_avg_byproduct) or None
    s1[(ROW_BY, COL_S_CUM)]      = summary.ytd_cum_byproduct or None

    s1[(ROW_WS, COL_S_PQTY)]     = summary.total_prev_waste_qty or None
    s1[(ROW_WS, COL_S_PAMT)]     = summary.total_prev_waste or None
    s1[(ROW_WS, COL_S_QTY)]      = ws_curr_qty or None
    s1[(ROW_WS, COL_S_AMT)]      = summary.total_current_waste or None
    s1[(ROW_WS, COL_S_PREV_AVG)] = (round(summary.prev_year_avg_waste)
                                    if summary.prev_year_avg_waste else None)
    s1[(ROW_WS, COL_S_AVG)]      = round(summary.ytd_avg_waste) or None
    s1[(ROW_WS, COL_S_CUM)]      = summary.ytd_cum_waste or None

    net_curr     = summary.total_current_byproduct + summary.total_current_waste
    net_prev     = summary.total_prev_byproduct    + summary.total_prev_waste
    net_avg      = summary.ytd_avg_byproduct       + summary.ytd_avg_waste
    net_cum      = summary.ytd_cum_byproduct       + summary.ytd_cum_waste
    net_prev_avg = (round(summary.prev_year_avg_byproduct + summary.prev_year_avg_waste)
                    if summary.prev_year_avg_byproduct is not None else None)

    s1[(ROW_NE, COL_S_PAMT)]     = net_prev or None
    s1[(ROW_NE, COL_S_AMT)]      = net_curr or None
    s1[(ROW_NE, COL_S_PREV_AVG)] = net_prev_avg
    s1[(ROW_NE, COL_S_AVG)]      = round(net_avg) or None
    s1[(ROW_NE, COL_S_CUM)]      = net_cum or None

    # ── 3-5. 세부현황 초기화 ─────────────────────────────────────────────────
    for r in range(BY_DATA_START, BY_TOTAL_ROW + 1):
        for c in CLEAR_DETAIL_COLS:
            s1[(r, c)] = None
    for r in range(WS_DATA_START, WS_TOTAL_ROW + 1):
        for c in CLEAR_DETAIL_COLS:
            s1[(r, c)] = None

    # ── 3-6. 세부현황 데이터 ─────────────────────────────────────────────────
    _fill_section_updates(s1, summary.byproducts, by_mapping, by_subtotals, BY_TOTAL_ROW)
    _fill_section_updates(s1, summary.wastes,     ws_mapping, ws_subtotals, WS_TOTAL_ROW)

    # ── 3-7. 그래프 시트 데이터 ──────────────────────────────────────────────
    # 상단 표 (rows 17-21): 원본 kg / 원 단위  → 차트 시리즈 참조 대상
    # 하단 표 (rows 24-28): 수량 ton(÷1000, 소수2자리) / 금액 백만원(÷1,000,000)
    def _chart_rows(s2, col, by_qty, by_amt, ws_qty, ws_amt):
        net = by_amt + ws_amt
        # 상단: 원본
        s2[(17, col)] = by_qty or None
        s2[(18, col)] = by_amt or None
        s2[(19, col)] = ws_qty or None
        s2[(20, col)] = ws_amt or None
        s2[(21, col)] = net or None
        # 하단: 단위 변환
        s2[(24, col)] = round(by_qty / 1000, 2) if by_qty else None
        s2[(25, col)] = round(by_amt / 1_000_000, 2) if by_amt else None
        s2[(26, col)] = round(ws_qty / 1000, 2) if ws_qty else None
        s2[(27, col)] = round(ws_amt / 1_000_000, 2) if ws_amt else None
        s2[(28, col)] = round(net / 1_000_000, 2) if net else None

    if trend_data and year_col_map:
        for m, d in trend_data.items():
            col = year_col_map.get(m)
            if not col:
                continue
            _chart_rows(s2, col, d['by_qty'], d['by_amt'], d['ws_qty'], d['ws_amt'])
    elif year_col_map and (chart_col := year_col_map.get(month)):
        by_qty  = sum(r.current_quantity or 0 for r in summary.byproducts)
        ws_qty  = sum(r.current_quantity or 0 for r in summary.wastes)
        _chart_rows(s2, chart_col, by_qty, summary.total_current_byproduct,
                    ws_qty, summary.total_current_waste)

    # ── 4. xlsx 직접 패치 ─────────────────────────────────────────────────────
    _patch_xlsx(output_path, {
        'xl/worksheets/sheet1.xml': s1,
        'xl/worksheets/sheet2.xml': s2,
    })

    return output_path
