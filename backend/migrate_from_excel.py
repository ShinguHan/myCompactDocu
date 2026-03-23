"""
엑셀 원본 파일에서 초기 데이터를 DB로 마이그레이션하는 1회성 스크립트.

실행 방법:
    cd backend
    python migrate_from_excel.py

대상 파일: ref/(찐) 26년_3월_부산물 매각 및 폐기물 처리현황.xlsx
"""
import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from datetime import date as date_type
from database import engine, SessionLocal, Base
import models

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "ref",
    "(찐) 26년_3월_부산물 매각 및 폐기물 처리현황.xlsx")


def find_sheet(wb, keyword):
    for name in wb.sheetnames:
        if keyword in name:
            return wb[name]
    return None


def load_master_data(db, ws_master):
    """
    '폐기물, 부산물 처리단가, 업체현황(목록표)' 시트에서
    품목, 업체, 계약단가, 품목-업체 연결 데이터를 로드.
    """
    print("\n[1] 기준정보 (목록표) 로드 중...")

    # 왼쪽 테이블: row3부터, 품목/단가/단위/업체1~4
    items_created = 0
    companies_created = 0
    links_created = 0
    contracts_created = 0

    # 오른쪽 테이블(J/K열)에서 폐기물 분류 목록 추출 (category 판별용)
    waste_categories = set()
    for row in ws_master.iter_rows(min_row=3, min_col=10, max_col=10, values_only=True):
        if row[0]:
            waste_categories.add(str(row[0]).strip())

    for row in ws_master.iter_rows(min_row=3, values_only=True):
        item_name = row[1]
        if not item_name:
            continue
        item_name = str(item_name).strip()
        unit_price_raw = row[2]
        unit = str(row[3]).strip() if row[3] else "원/kg"

        # 단가가 '변동'이거나 없는 경우
        if unit_price_raw is None or str(unit_price_raw).strip() in ('', '변동'):
            unit_price = 0.0
            unit_type = "fixed"
        else:
            try:
                unit_price = float(unit_price_raw)
                unit_type = "per_unit"
            except (ValueError, TypeError):
                unit_price = 0.0
                unit_type = "fixed"

        # 카테고리: 오른쪽 테이블의 폐기물 분류에 있으면 폐기물, 아니면 부산물
        # 단가가 음수면 폐기물 처리비용
        if unit_price < 0:
            category = "폐기물"
        elif item_name in waste_categories:
            category = "폐기물"
        else:
            category = "부산물"

        # 품목 생성
        item = db.query(models.Item).filter(models.Item.name == item_name).first()
        if not item:
            item = models.Item(
                name=item_name,
                report_name=None,
                unit=unit,
                category=category,
            )
            db.add(item)
            db.flush()
            items_created += 1

        # 업체1~4 (col index 4~7)
        for sort_order, company_name_raw in enumerate(row[4:8], start=1):
            if not company_name_raw:
                continue
            company_name = str(company_name_raw).strip()

            # 업체 생성
            company = db.query(models.Company).filter(models.Company.name == company_name).first()
            if not company:
                company = models.Company(name=company_name)
                db.add(company)
                db.flush()
                companies_created += 1

            # 품목-업체 연결
            link = db.query(models.ItemCompany).filter(
                models.ItemCompany.item_id == item.id,
                models.ItemCompany.company_id == company.id,
            ).first()
            if not link:
                link = models.ItemCompany(
                    item_id=item.id, company_id=company.id, sort_order=sort_order
                )
                db.add(link)
                links_created += 1

            # 계약단가 (sort_order==1인 주 업체만)
            if sort_order == 1 and unit_price != 0:
                contract = db.query(models.Contract).filter(
                    models.Contract.item_id == item.id,
                    models.Contract.company_id == company.id,
                ).first()
                if not contract:
                    contract = models.Contract(
                        item_id=item.id,
                        company_id=company.id,
                        unit_price=abs(unit_price),
                        unit_type=unit_type,
                        effective_date=date_type(2026, 1, 1),
                    )
                    db.add(contract)
                    contracts_created += 1

    db.commit()
    print(f"  품목: {items_created}개, 업체: {companies_created}개, "
          f"연결: {links_created}개, 계약: {contracts_created}개 생성")


def load_transactions(db, ws_ledger):
    """
    '폐기물, 부산물 입출고대장' 시트에서 거래 데이터 로드.
    row2: 헤더 (일자, 품목, 처리업체, 데이터변환, 계약단가, 단위, 처리량, 금액, 비고)
    """
    print("\n[2] 입출고대장 거래 데이터 로드 중...")

    # 미리 item/company 이름→ID 매핑
    item_map = {i.name: i.id for i in db.query(models.Item).all()}
    company_map = {c.name: c.id for c in db.query(models.Company).all()}

    created = 0
    skipped = 0
    unknown = set()

    for row in ws_ledger.iter_rows(min_row=3, values_only=True):
        tx_date = row[1]
        item_name = row[2]
        company_name = row[3]
        unit_price_raw = row[5]
        quantity_raw = row[7]
        total_amount_raw = row[8]
        note = row[9]

        if not tx_date or not item_name or not company_name:
            continue

        # 날짜 변환
        if hasattr(tx_date, 'date'):
            tx_date = tx_date.date()
        elif isinstance(tx_date, str):
            try:
                from datetime import datetime
                tx_date = datetime.strptime(tx_date, '%Y-%m-%d').date()
            except Exception:
                continue

        item_name = str(item_name).strip()
        company_name = str(company_name).strip()

        item_id = item_map.get(item_name)
        company_id = company_map.get(company_name)

        # 미등록 품목/업체 → 자동 생성
        if not item_id:
            unknown.add(f"품목:{item_name}")
            item = models.Item(name=item_name, unit="원/kg", category="폐기물")
            db.add(item)
            db.flush()
            item_map[item_name] = item.id
            item_id = item.id

        if not company_id:
            unknown.add(f"업체:{company_name}")
            company = models.Company(name=company_name)
            db.add(company)
            db.flush()
            company_map[company_name] = company.id
            company_id = company.id

        # 중복 체크
        exists = db.query(models.Transaction).filter(
            models.Transaction.date == tx_date,
            models.Transaction.item_id == item_id,
            models.Transaction.company_id == company_id,
        ).first()
        if exists:
            skipped += 1
            continue

        try:
            unit_price = float(unit_price_raw or 0)
            quantity = float(quantity_raw or 0)
            total_amount = float(total_amount_raw or 0)
        except (ValueError, TypeError):
            skipped += 1
            continue

        tx = models.Transaction(
            date=tx_date,
            item_id=item_id,
            company_id=company_id,
            quantity=quantity,
            unit_price=unit_price,
            total_amount=total_amount,
            note=str(note) if note else None,
        )
        db.add(tx)
        created += 1

        if created % 50 == 0:
            db.flush()

    db.commit()
    print(f"  거래: {created}건 생성, {skipped}건 건너뜀")
    if unknown:
        print(f"  자동 생성된 미등록 항목: {', '.join(sorted(unknown))}")


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] 파일을 찾을 수 없습니다: {EXCEL_PATH}")
        sys.exit(1)

    print(f"파일 로드: {os.path.basename(EXCEL_PATH)}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    ws_master = find_sheet(wb, "목록표")
    ws_ledger = find_sheet(wb, "입출고대장")

    if not ws_master:
        print("[ERROR] '목록표' 시트를 찾을 수 없습니다")
        sys.exit(1)
    if not ws_ledger:
        print("[ERROR] '입출고대장' 시트를 찾을 수 없습니다")
        sys.exit(1)

    # DB 테이블 생성
    Base.metadata.create_all(bind=engine)

    db = SessionLocal()
    try:
        load_master_data(db, ws_master)
        load_transactions(db, ws_ledger)
        print("\n✅ 마이그레이션 완료!")
    except Exception as e:
        db.rollback()
        print(f"\n[ERROR] {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
