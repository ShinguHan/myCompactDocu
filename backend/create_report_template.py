"""
기존 참조 엑셀 파일에서 데이터만 비운 보고서 템플릿 생성
한 번만 실행하면 됨
"""
import openpyxl
import shutil
import os

SRC = os.path.join(os.path.dirname(__file__), '..', 'ref',
                   '(찐) 26년_3월_부산물 매각 및 폐기물 처리현황.xlsx')
DST = os.path.join(os.path.dirname(__file__), 'templates', 'report_template.xlsx')

os.makedirs(os.path.dirname(DST), exist_ok=True)
shutil.copy(SRC, DST)
print(f'Copied: {DST}')

wb = openpyxl.load_workbook(DST)
ws = wb.worksheets[0]  # 메인 시트

def safe_clear_num(ws, row, col):
    """숫자·텍스트 데이터만 비움 (병합셀 무시, '-' 유지)"""
    try:
        cell = ws.cell(row, col)
        if isinstance(cell.value, (int, float)):
            cell.value = None
        elif isinstance(cell.value, str) and cell.value.strip() not in ('-',):
            cell.value = None
    except AttributeError:
        pass

# ── 1. 요약 섹션 (rows 8-10) ─────────────────────────────────────────────────
for r in range(8, 11):
    for c in [4, 6, 9, 11, 14, 17, 20]:   # D F I K N Q T
        safe_clear_num(ws, r, c)

# ── 2. 부산물 세부 (rows 16-36 데이터 + row 37 합계) ─────────────────────────
for r in range(16, 38):
    for c in [12, 14, 17, 20]:            # L N Q T
        safe_clear_num(ws, r, c)

# ── 3. 폐기물 세부 (rows 42-55 데이터 + row 56 합계) ─────────────────────────
for r in range(42, 57):
    for c in [12, 14, 17, 20]:
        safe_clear_num(ws, r, c)

# ── 4. 연도/월 (B2, G2) ───────────────────────────────────────────────────────
safe_clear_num(ws, 2, 2)
safe_clear_num(ws, 2, 7)

wb.save(DST)
print('템플릿 저장 완료')

# ── 검증 ──────────────────────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(DST)
ws2 = wb2.worksheets[0]
print(f'연도={ws2["B2"].value}  월={ws2["G2"].value}')
print(f'D8(수량)={ws2["D8"].value}  F8(금액)={ws2["F8"].value}  K8(전월)={ws2["K8"].value}')
print(f'L16(수량)={ws2["L16"].value}  N16(금액)={ws2["N16"].value}')
print(f'업체명 유지 B16={ws2["B16"].value}')
print(f'품명 유지 F16={ws2["F16"].value}')
print(f'단가 유지 J16={ws2["J16"].value}')
print(f'시트목록: {wb2.sheetnames}')
