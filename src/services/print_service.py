import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
import shutil
import re
import os

class PrintService:
    # Template layout constants
    TEMPLATE_COLS = 17   # A(1) to Q(17)
    TEMPLATE_ROWS = 81   # rows 1 to 81
    DATE_ROW = 10
    DATE_COL = 11        # K
    ITEM_NAME_ROW = 18
    ITEM_NAME_COL = 1    # A
    ITEM_QTY_COL = 12    # L
    IMAGE_START_ROW = 36 # Receipt image area start
    IMAGE_END_ROW = 81   # Receipt image area end

    def __init__(self, template_path, images_dir=None):
        self.template_path = template_path
        self.images_dir = images_dir

    def generate_exit_pass(self, transaction_data, output_path):
        """
        Generates an Exit Pass Excel file.
        Each item gets its own template copy placed side by side horizontally.

        transaction_data: dict containing:
            - date: datetime.date
            - company_name: str
            - items: list of dicts {'name': str, 'quantity': float, 'amount': float}
        """
        shutil.copy(self.template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
        ws = wb['반출증']

        items = transaction_data.get('items', [])
        date = transaction_data['date']

        if not items:
            wb.save(output_path)
            return output_path

        company_name = transaction_data.get('company_name', '')

        # First item fills the existing template (col_offset=0)
        self._fill_item(ws, date, items[0], col_offset=0, company_name=company_name)

        # Each additional item: copy template block to the right, then fill
        for i, item in enumerate(items[1:], start=1):
            col_offset = i * self.TEMPLATE_COLS
            self._copy_template_block(ws, col_offset)
            self._fill_item(ws, date, item, col_offset=col_offset, company_name=company_name)

        # Set print area and page breaks based on item count
        self._set_print_layout(ws, len(items))

        # TODO: Insert receipt images into rows 36-81 for each item
        # if self.images_dir:
        #     for i, item in enumerate(items):
        #         img_path = self._find_image(date, transaction_data['company_name'], item['name'])
        #         if img_path:
        #             self._insert_image(ws, img_path, row=self.IMAGE_START_ROW,
        #                                col=1 + i * self.TEMPLATE_COLS)

        wb.save(output_path)
        return output_path

    def _fill_item(self, ws, date, item, col_offset, company_name=''):
        """Fill date, company, item name, and quantity at the given column offset."""
        # C8: clear broken #REF! serial number formula
        ws.cell(row=8, column=3 + col_offset).value = None

        # K10: date
        ws.cell(row=self.DATE_ROW, column=self.DATE_COL + col_offset).value = date

        # A12: company name (override VLOOKUP with actual value)
        ws.cell(row=12, column=1 + col_offset).value = company_name

        # A18: item name, L18: quantity
        ws.cell(row=self.ITEM_NAME_ROW, column=self.ITEM_NAME_COL + col_offset).value = item['name']
        ws.cell(row=self.ITEM_NAME_ROW, column=self.ITEM_QTY_COL + col_offset).value = item['quantity']

    def _set_print_layout(self, ws, item_count):
        """
        Set print area and column page breaks based on item count.
        Each item = 2 A4 pages (rows 1-35: 반출증, rows 36-81: 영수증 사진).
        Column page breaks are inserted every TEMPLATE_COLS columns.
        """
        # Expand print area to cover all item copies
        last_col = get_column_letter(item_count * self.TEMPLATE_COLS)
        ws.print_area = f"$A$1:${last_col}${self.TEMPLATE_ROWS}"

        # Add column page breaks between each item copy (after col 17, 34, 51, ...)
        from openpyxl.worksheet.pagebreak import Break
        for i in range(1, item_count):
            break_col = i * self.TEMPLATE_COLS  # after Q(17), AH(34), ...
            ws.col_breaks.append(Break(id=break_col))

    def _copy_template_block(self, ws, col_offset):
        """Copy A1:Q81 template to a new position shifted by col_offset columns."""

        # 1. Copy cell values and styles
        for row in range(1, self.TEMPLATE_ROWS + 1):
            for col in range(1, self.TEMPLATE_COLS + 1):
                src = ws.cell(row=row, column=col)
                dst = ws.cell(row=row, column=col + col_offset)

                if src.value is not None:
                    if isinstance(src.value, str) and src.value.startswith('='):
                        dst.value = self._shift_formula(src.value, col_offset)
                    else:
                        dst.value = src.value

                if src.has_style:
                    dst.font = copy(src.font)
                    dst.border = copy(src.border)
                    dst.fill = copy(src.fill)
                    dst.alignment = copy(src.alignment)
                    dst.number_format = src.number_format

        # 2. Recreate merged cells at new position
        new_merges = []
        for mr in list(ws.merged_cells.ranges):
            if (1 <= mr.min_col <= self.TEMPLATE_COLS and
                    1 <= mr.max_col <= self.TEMPLATE_COLS and
                    1 <= mr.min_row <= self.TEMPLATE_ROWS and
                    1 <= mr.max_row <= self.TEMPLATE_ROWS):
                min_col_l = get_column_letter(mr.min_col + col_offset)
                max_col_l = get_column_letter(mr.max_col + col_offset)
                new_merges.append(f"{min_col_l}{mr.min_row}:{max_col_l}{mr.max_row}")

        for m in new_merges:
            ws.merge_cells(m)

        # 3. Copy column widths
        # Template stores column widths as ranges (e.g. min=1, max=17, width=5.14).
        # Must use range-aware copy, not letter-by-letter lookup.
        for key, src_dim in list(ws.column_dimensions.items()):
            min_col = src_dim.min or column_index_from_string(key)
            max_col = src_dim.max or column_index_from_string(key)

            # Only process dimensions overlapping the template range (1~TEMPLATE_COLS)
            if min_col > self.TEMPLATE_COLS or max_col < 1:
                continue

            eff_min = max(min_col, 1)
            eff_max = min(max_col, self.TEMPLATE_COLS)

            for col in range(eff_min, eff_max + 1):
                dst_letter = get_column_letter(col + col_offset)
                ws.column_dimensions[dst_letter].width = src_dim.width

    def _shift_formula(self, formula, col_offset):
        """
        Shift internal cell column references in a formula by col_offset.
        External sheet references (e.g. '시트명'!$B$1) are left unchanged.
        """
        if '#REF!' in formula:
            return formula  # keep broken refs as-is

        # Split formula into external-ref parts and non-external parts
        # External refs look like: 'Sheet name'!$A$1:$Z$99 or SheetName!A1
        ext_ref_pattern = r"('[^']+'![^\s,();]+|[A-Za-z_]\w*![^\s,();]+)"
        parts = re.split(ext_ref_pattern, formula)

        result = []
        for part in parts:
            if '!' in part:
                result.append(part)  # external ref — do not shift
            else:
                result.append(self._shift_internal_refs(part, col_offset))

        return ''.join(result)

    def _shift_internal_refs(self, text, col_offset):
        """Shift all cell column references within a formula fragment."""
        def replacer(match):
            dollar1 = match.group(1)   # optional $ before col
            col_str = match.group(2)   # column letter(s)
            dollar2 = match.group(3)   # optional $ before row
            row_str = match.group(4)   # row digits

            try:
                col_num = column_index_from_string(col_str)
                new_col = get_column_letter(col_num + col_offset)
                return dollar1 + new_col + dollar2 + row_str
            except Exception:
                return match.group(0)

        # Match cell refs not preceded by ! or word char (to avoid matching inside names)
        pattern = r'(?<![!\w])(\$?)([A-Z]{1,3})(\$?)(\d+)'
        return re.sub(pattern, replacer, text)

    # ─────────────────────────────────────────────
    # Image management (for future use)
    # ─────────────────────────────────────────────

    def get_image_path(self, images_dir, date, company_name, item_name):
        """
        Returns expected image path for a given date/company/item.
        Looks for jpg, jpeg, png files.
        """
        base = f"{date}_{company_name}_{item_name}".replace("/", "_").replace(" ", "_")
        for ext in ('jpg', 'jpeg', 'png'):
            path = os.path.join(images_dir, str(date)[:7], f"{base}.{ext}")
            if os.path.exists(path):
                return path
        return None
